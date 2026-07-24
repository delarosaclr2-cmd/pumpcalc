import openpyxl
import csv
from pathlib import Path

FILE_PATH = Path(r"C:\PUMPCALC\original\KEETP-60-DM-008 - HOJA DE ESPECIFICACIÓN BOMBA 005PU001 REV C (1).xlsm")
REPORTS_DIR = Path(r"C:\PUMPCALC\reports")

def audit_named_ranges():
    wb = openpyxl.load_workbook(FILE_PATH, read_only=False, keep_vba=True, data_only=False)
    
    audit_data = []
    
    # Get all named ranges
    for nr in wb.defined_names.values():
        name = nr.name
        ref = nr.attr_text if nr.attr_text else ""
        comment = nr.comment if hasattr(nr, 'comment') and nr.comment else ""
        
        # Parse the reference to get sheet and range
        sheet_name = None
        cell_range = None
        has_ref_error = "#REF!" in ref
        
        # Try to determine scope
        is_global = True
        if hasattr(nr, 'localSheetId') and nr.localSheetId is not None:
            is_global = False
        
        # Find which formulas use this named range
        usage_count = 0
        usage_cells = []
        
        for ws_name in wb.sheetnames:
            ws = wb[ws_name]
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column, values_only=False):
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        if name in cell.value:
                            usage_count += 1
                            if len(usage_cells) < 20:
                                usage_cells.append(f"{ws_name}!{cell.coordinate}")
        
        # Check data validations
        dv_usage = 0
        dv_cells = []
        for ws_name in wb.sheetnames:
            ws = wb[ws_name]
            if ws.data_validations:
                for dv in ws.data_validations.dataValidation:
                    if dv.formula1 and name in str(dv.formula1):
                        dv_usage += 1
                    if dv.formula2 and name in str(dv.formula2):
                        dv_usage += 1
        
        # Determine possible replacement for LISTA
        possible_replacement = ""
        if name == "LISTA":
            possible_replacement = "Likely a deleted table or data validation list source"
        
        audit_data.append({
            "name": name,
            "scope": "Global" if is_global else f"Local (sheetId={nr.localSheetId})",
            "reference": ref,
            "has_ref_error": has_ref_error,
            "sheet_associated": sheet_name,
            "cell_range": cell_range,
            "used_in_formulas": usage_count,
            "formula_cells": "; ".join(usage_cells),
            "used_in_data_validation": dv_usage,
            "possible_replacement": possible_replacement,
            "comment": comment
        })
    
    # Save CSV
    csv_path = REPORTS_DIR / "named_ranges_audit.csv"
    with open(csv_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=audit_data[0].keys())
        writer.writeheader()
        writer.writerows(audit_data)
    
    # Save Markdown
    md_path = REPORTS_DIR / "named_ranges_audit.md"
    with open(md_path, 'w', encoding='utf-8') as f:
        f.write("# Named Ranges Audit\n\n")
        f.write(f"**File:** {FILE_PATH.name}\n")
        f.write(f"**Total named ranges:** {len(audit_data)}\n\n")
        
        f.write("## Summary Table\n\n")
        f.write("| Name | Scope | Reference | #REF! | Formulas Using | DV Using | Notes |\n")
        f.write("|------|-------|-----------|-------|----------------|----------|-------|\n")
        for nr in audit_data:
            ref_short = nr['reference'][:60] + "..." if len(nr['reference']) > 60 else nr['reference']
            notes = "⚠️ BROKEN" if nr['has_ref_error'] else ""
            if nr['possible_replacement']:
                notes += f" - {nr['possible_replacement']}"
            f.write(f"| **{nr['name']}** | {nr['scope']} | `{ref_short}` | {'Yes' if nr['has_ref_error'] else 'No'} | {nr['used_in_formulas']} | {nr['used_in_data_validation']} | {notes} |\n")
        
        f.write("\n## Detailed Analysis\n\n")
        for nr in audit_data:
            f.write(f"### {nr['name']}\n")
            f.write(f"- **Scope:** {nr['scope']}\n")
            f.write(f"- **Reference:** `{nr['reference']}`\n")
            f.write(f"- **Contains #REF!:** {'YES ⚠️' if nr['has_ref_error'] else 'No'}\n")
            f.write(f"- **Used in formulas:** {nr['used_in_formulas']} cells\n")
            if nr['formula_cells']:
                f.write(f"  - Cells: {nr['formula_cells']}\n")
            f.write(f"- **Used in data validation:** {nr['used_in_data_validation']}\n")
            if nr['possible_replacement']:
                f.write(f"- **Note:** {nr['possible_replacement']}\n")
            f.write("\n")
        
        # Special focus on LISTA
        lista = [nr for nr in audit_data if nr['name'] == 'LISTA']
        if lista:
            nr = lista[0]
            f.write("## Special Analysis: LISTA (Broken Reference)\n\n")
            f.write(f"- **Reference:** `{nr['reference']}`\n")
            f.write(f"- **Used in formulas:** {nr['used_in_formulas']}\n")
            f.write(f"- **Used in data validation:** {nr['used_in_data_validation']}\n")
            f.write(f"- **Assessment:** ")
            if nr['used_in_formulas'] == 0 and nr['used_in_data_validation'] == 0:
                f.write("**Completely unused** - safe to delete or ignore\n")
            elif nr['used_in_formulas'] > 0:
                f.write(f"**USED in {nr['used_in_formulas']} formula(s)** - MUST be fixed\n")
            else:
                f.write("Used only in data validation - affects dropdown lists\n")
            f.write("\n")
    
    return audit_data

if __name__ == "__main__":
    audit_data = audit_named_ranges()
    print(f"Audited {len(audit_data)} named ranges")
    for nr in audit_data:
        if nr['name'] == 'LISTA':
            print(f"LISTA: #REF!={nr['has_ref_error']}, used_in_formulas={nr['used_in_formulas']}, used_in_dv={nr['used_in_data_validation']}")