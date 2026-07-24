import openpyxl
import csv
import re
from pathlib import Path

FILE_PATH = Path(r"C:\PUMPCALC\original\KEETP-60-DM-008 - HOJA DE ESPECIFICACIÓN BOMBA 005PU001 REV C (1).xlsm")
REPORTS_DIR = Path(r"C:\PUMPCALC\reports")

def extract_all_formulas():
    wb = openpyxl.load_workbook(FILE_PATH, read_only=False, keep_vba=True, data_only=False)
    
    formulas = []
    constants = []
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column, values_only=False):
            for cell in row:
                if cell.value is not None:
                    if isinstance(cell.value, str) and cell.value.startswith('='):
                        # It's a formula
                        formula = cell.value
                        
                        # Find named ranges used
                        named_ranges_used = []
                        for nr_name in wb.defined_names:
                            if nr_name in formula:
                                named_ranges_used.append(nr_name)
                        
                        # Find sheet references
                        sheet_refs = re.findall(r"'([^']+)'!", formula)
                        
                        # Find numeric constants in formula
                        numeric_constants = re.findall(r'(?<![A-Z])(\d+\.?\d*)(?![A-Z])', formula)
                        numeric_constants = [float(c) for c in numeric_constants if float(c) not in [0, 1]]
                        
                        # Extract function names
                        functions = re.findall(r'([A-ZÁÉÍÓÚÑ]+)\s*\(', formula)
                        
                        formulas.append({
                            "sheet": sheet_name,
                            "cell": cell.coordinate,
                            "formula_a1": formula,
                            "formula_r1c1": "",  # Will convert if needed
                            "cached_value": cell.value if cell.data_type == 'n' else (cell.value if cell.data_type != 'f' else "FORMULA"),
                            "number_format": cell.number_format,
                            "data_type": cell.data_type,
                            "named_ranges_used": "; ".join(named_ranges_used),
                            "sheet_references": "; ".join(set(sheet_refs)),
                            "numeric_constants": "; ".join(str(c) for c in numeric_constants),
                            "functions_used": "; ".join(set(functions)),
                            "possible_meaning": "",
                            "input_units": "",
                            "output_units": "",
                            "confidence": ""
                        })
                    elif cell.data_type in ('n', 's') and not isinstance(cell.value, str):
                        # Constant value
                        constants.append({
                            "sheet": sheet_name,
                            "cell": cell.coordinate,
                            "value": cell.value,
                            "data_type": cell.data_type,
                            "number_format": cell.number_format,
                            "is_formula": False
                        })
    
    # Save formulas CSV
    csv_path = REPORTS_DIR / "formula_inventory.csv"
    with open(csv_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=formulas[0].keys() if formulas else [])
        writer.writeheader()
        writer.writerows(formulas)
    
    # Save constants CSV
    const_csv_path = REPORTS_DIR / "constants_inventory.csv"
    with open(const_csv_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=constants[0].keys() if constants else [])
        writer.writeheader()
        writer.writerows(constants)
    
    # Generate Markdown report
    md_path = REPORTS_DIR / "formula_inventory.md"
    with open(md_path, 'w', encoding='utf-8') as f:
        f.write("# Formula Inventory\n\n")
        f.write(f"**Total formulas:** {len(formulas)}\n")
        f.write(f"**Total constants:** {len(constants)}\n\n")
        
        # Group by sheet
        by_sheet = {}
        for frm in formulas:
            sheet = frm['sheet']
            if sheet not in by_sheet:
                by_sheet[sheet] = []
            by_sheet[sheet].append(frm)
        
        for sheet_name, sheet_formulas in by_sheet.items():
            f.write(f"## Sheet: `{sheet_name}` ({len(sheet_formulas)} formulas)\n\n")
            f.write("| Cell | Formula | Named Ranges | Sheet Refs | Constants | Functions | Format |\n")
            f.write("|------|---------|--------------|------------|-----------|-----------|--------|\n")
            for frm in sheet_formulas:
                formula_short = frm['formula_a1'][:80] + "..." if len(frm['formula_a1']) > 80 else frm['formula_a1']
                nr = frm['named_ranges_used'][:50] + "..." if len(frm['named_ranges_used']) > 50 else frm['named_ranges_used']
                sr = frm['sheet_references'][:50] + "..." if len(frm['sheet_references']) > 50 else frm['sheet_references']
                nc = frm['numeric_constants'][:50] + "..." if len(frm['numeric_constants']) > 50 else frm['numeric_constants']
                fn = frm['functions_used'][:50] + "..." if len(frm['functions_used']) > 50 else frm['functions_used']
                f.write(f"| {frm['cell']} | `{formula_short}` | {nr} | {sr} | {nc} | {fn} | {frm['number_format']} |\n")
            f.write("\n")
    
    return formulas, constants

if __name__ == "__main__":
    formulas, constants = extract_all_formulas()
    print(f"Extracted {len(formulas)} formulas and {len(constants)} constants")