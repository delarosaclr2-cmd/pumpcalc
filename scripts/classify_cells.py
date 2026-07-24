import openpyxl
import csv
from pathlib import Path

FILE_PATH = Path(r"C:\PUMPCALC\original\KEETP-60-DM-008 - HOJA DE ESPECIFICACIÓN BOMBA 005PU001 REV C (1).xlsm")
REPORTS_DIR = Path(r"C:\PUMPCALC\reports")

def get_fill_color(cell):
    """Safely extract fill color from cell"""
    try:
        if cell.fill and cell.fill.start_color and cell.fill.start_color.rgb:
            return str(cell.fill.start_color.rgb)
    except (AttributeError, TypeError):
        pass  # GradientFill or other fill types
    return ""

def classify_cells():
    wb = openpyxl.load_workbook(FILE_PATH, read_only=False, keep_vba=True, data_only=False)
    
    classification_data = []
    
    # Key sheets to analyze
    key_sheets = [
        'CAIDA PRESION DE TUBERIA',
        'CALCULO DE BOMBA',
        'TABLA DE ACCESORIOS DESCARGA',
        'TABLA DE ACCESORIOS SUCCION',
        'RAMALES',
        '005PU001',
        'RESUMEN PARA PDF',
        'REPORTE GENERAL',
        'ESPECIFICACIÓN DE TUBERIA',
        'VELOCIDADES RECOMENDADAS'
    ]
    
    for sheet_name in key_sheets:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column, values_only=False):
            for cell in row:
                if cell.value is not None or cell.comment or cell.data_type != 'n':
                    # Determine classification
                    classification = classify_cell(cell, ws, wb)
                    
                    classification_data.append({
                        "sheet": sheet_name,
                        "cell": cell.coordinate,
                        "value": str(cell.value)[:100] if cell.value else "",
                        "data_type": cell.data_type,
                        "number_format": cell.number_format,
                        "is_formula": isinstance(cell.value, str) and cell.value.startswith('='),
                        "is_constant": cell.value is not None and not (isinstance(cell.value, str) and cell.value.startswith('=')),
                        "has_comment": cell.comment is not None,
                        "comment_text": cell.comment.text[:100] if cell.comment and cell.comment.text else "",
                        "classification": classification,
                        "protection_locked": cell.protection.locked if cell.protection else True,
                        "protection_hidden": cell.protection.hidden if cell.protection else False,
                        "fill_color": get_fill_color(cell),
                        "font_color": str(cell.font.color.rgb) if cell.font and cell.font.color and cell.font.color.rgb else "",
                        "border_style": str(cell.border) if cell.border else ""
                    })
    
    # Save CSV
    csv_path = REPORTS_DIR / "cell_classification.csv"
    with open(csv_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=classification_data[0].keys())
        writer.writeheader()
        writer.writerows(classification_data)
    
    # Save Markdown
    md_path = REPORTS_DIR / "input_map.md"
    with open(md_path, 'w', encoding='utf-8') as f:
        f.write("# Cell Classification & Input Map\n\n")
        f.write(f"**Total cells classified:** {len(classification_data)}\n\n")
        
        # Group by classification
        by_class = {}
        for c in classification_data:
            cls = c['classification']
            if cls not in by_class:
                by_class[cls] = []
            by_class[cls].append(c)
        
        for cls in sorted(by_class.keys()):
            f.write(f"\n## {cls} ({len(by_class[cls])} cells)\n\n")
            f.write("| Sheet | Cell | Value/Formula | Format | Locked | Fill |\n")
            f.write("|-------|------|---------------|--------|--------|------|\n")
            for c in by_class[cls][:50]:  # Limit to 50 per class
                val = c['value']
                if len(val) > 60:
                    val = val[:60] + "..."
                f.write(f"| {c['sheet']} | {c['cell']} | `{val}` | {c['number_format']} | {c['protection_locked']} | {c['fill_color']} |\n")
    
    return classification_data

def classify_cell(cell, ws, wb):
    """Classify a cell based on its properties and context"""
    coord = cell.coordinate
    val = cell.value
    
    # Check if it's a formula
    is_formula = isinstance(val, str) and val.startswith('=')
    
    # Check if unlocked (potential input)
    is_unlocked = cell.protection and not cell.protection.locked
    
    # Check for data validation
    has_dv = False
    if ws.data_validations:
        for dv in ws.data_validations.dataValidation:
            if coord in dv:
                has_dv = True
                break
    
    # Classification logic
    row = cell.row
    col = cell.column
    
    if is_formula:
        # Check if it's a final output (referenced by many, or in output sheets)
        if ws.title in ['005PU001', 'RESUMEN PARA PDF', 'REPORTE GENERAL']:
            return "FINAL_OUTPUT"
        # Check if it's a lookup/reference
        if 'VLOOKUP' in val or 'HLOOKUP' in val or 'INDEX' in val or 'MATCH' in val:
            return "LOOKUP"
        # Check if it references manufacturer data sheets
        if any(s in val for s in ['VELOCIDADES RECOMENDADAS', 'OUTPIPES', 'INPIPE', 'FRICCION', 'DIAMETRO', 'RUGOSIDAD']):
            return "LOOKUP"
        return "INTERMEDIATE_CALCULATION"
    
    # Constants
    if val is not None and not is_formula:
        # Check if it's a design/process input
        if has_dv or is_unlocked:
            return "USER_INPUT"
        
        # Check for typical input indicators
        cell_str = str(val).upper()
        if any(kw in cell_str for kw in ['FLUIDO', 'CAUDAL', 'FLUJO', 'PRESION', 'TEMPERATURA', 'DENSIDAD', 'VISCOSIDAD', 'DIAMETRO', 'LONGITUD', 'RUGOSIDAD', 'TAG', 'NOMBRE', 'PROYECTO']):
            return "DESIGN_INPUT"
        
        # Check if it's in a known data table sheet
        if ws.title == 'VELOCIDADES RECOMENDADAS':
            if row <= 4:  # Headers
                return "LOOKUP"
            return "MANUFACTURER_DATA"
        
        if ws.title == 'ESPECIFICACIÓN DE TUBERIA':
            return "PIPE_DATA"
        
        if ws.title == 'REGISTROS':
            return "MANUFACTURER_DATA"
        
        # Check if it's a numeric constant in calculation area
        if isinstance(val, (int, float)) and ws.title in ['CAIDA PRESION DE TUBERIA', 'CALCULO DE BOMBA', 'TABLA DE ACCESORIOS DESCARGA', 'TABLA DE ACCESORIOS SUCCION', 'RAMALES']:
            # Could be assumption or constant
            if val in [14.7, 2.31, 3960, 0.7456, 5252, 1700, 50.6, 2.3071, 32.4, 448.8309, 3.281, 0.3048, 3.7854]:
                return "CONSTANT"
            return "ASSUMPTION"
        
        return "UNKNOWN"
    
    return "DECORATIVE"

if __name__ == "__main__":
    data = classify_cells()
    print(f"Classified {len(data)} cells")