import openpyxl
import json
import hashlib
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook

FILE_PATH = Path(r"C:\PUMPCALC\original\KEETP-60-DM-008 - HOJA DE ESPECIFICACIÓN BOMBA 005PU001 REV C (1).xlsm")
WORKING_PATH = Path(r"C:\PUMPCALC\working\KEETP-60-DM-008 - HOJA DE ESPECIFICACIÓN BOMBA 005PU001 REV C (1).xlsm")
REPORTS_DIR = Path(r"C:\PUMPCALC\reports")

def compute_hashes(filepath):
    hashes = {}
    for algo in ['md5', 'sha256', 'sha1']:
        h = hashlib.new(algo)
        with open(filepath, 'rb') as f:
            for chunk in iter(lambda: f.read(8192), b''):
                h.update(chunk)
        hashes[algo] = h.hexdigest()
    return hashes

def inspect_workbook(filepath):
    wb = load_workbook(filepath, read_only=False, keep_vba=True, data_only=False)
    
    inventory = {
        "file_info": {
            "path": str(filepath),
            "size_bytes": filepath.stat().st_size,
            "extension": filepath.suffix,
            "hashes": compute_hashes(filepath),
            "inspection_time": datetime.now().isoformat(),
        },
        "workbook": {
            "sheet_names": wb.sheetnames,
            "sheet_count": len(wb.sheetnames),
            "active_sheet": wb.active.title if wb.active else None,
            "has_vba": wb.vba_archive is not None,
            "properties": {
                "creator": wb.properties.creator,
                "last_modified_by": wb.properties.lastModifiedBy,
                "created": str(wb.properties.created) if wb.properties.created else None,
                "modified": str(wb.properties.modified) if wb.properties.modified else None,
            }
        },
        "sheets": {}
    }
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_info = {
            "name": sheet_name,
            "title": ws.title,
            "sheet_state": ws.sheet_state,
            "dimensions": ws.dimensions,
            "max_row": ws.max_row,
            "max_column": ws.max_column,
            "sheet_format": {
                "default_row_height": ws.sheet_format.defaultRowHeight,
                "default_col_width": ws.sheet_format.defaultColWidth,
            } if ws.sheet_format else {},
            "page_setup": {
                "orientation": ws.page_setup.orientation,
                "paper_size": ws.page_setup.paperSize,
            } if ws.page_setup else {},
            "print_options": {
                "horizontal_centered": ws.print_options.horizontalCentered,
                "vertical_centered": ws.print_options.verticalCentered,
            } if ws.print_options else {},
            "merged_cells": [str(m) for m in ws.merged_cells.ranges],
            "conditional_formats": len(ws.conditional_formatting._cf_rules) if ws.conditional_formatting else 0,
            "data_validations": len(ws.data_validations.dataValidation) if ws.data_validations else 0,
            "tables": [t.name for t in ws.tables.values()] if ws.tables else [],
            "charts": len(ws._charts) if hasattr(ws, '_charts') else 0,
            "images": len(ws._images) if hasattr(ws, '_images') else 0,
            "drawings": len(ws._drawings) if hasattr(ws, '_drawings') else 0,
            "comments": len(ws._comments) if hasattr(ws, '_comments') and ws._comments else 0,
            "hyperlinks": len(ws._hyperlinks) if hasattr(ws, '_hyperlinks') and ws._hyperlinks else 0,
            "auto_filter": str(ws.auto_filter.ref) if ws.auto_filter and ws.auto_filter.ref else None,
            "freeze_panes": ws.freeze_panes,
            "print_area": ws.print_area,
            "print_titles": ws.print_titles,
            "sheet_properties": {
                "tab_color": str(ws.sheet_properties.tabColor.rgb) if ws.sheet_properties and ws.sheet_properties.tabColor and ws.sheet_properties.tabColor.rgb else None,
                "outline_summary_right": ws.sheet_properties.outlinePr.summaryRight if ws.sheet_properties and ws.sheet_properties.outlinePr else None,
                "outline_summary_below": ws.sheet_properties.outlinePr.summaryBelow if ws.sheet_properties and ws.sheet_properties.outlinePr else None,
                "outline_show_symbols": ws.sheet_properties.outlinePr.showOutlineSymbols if ws.sheet_properties and ws.sheet_properties.outlinePr else None,
            },
        }
        
        # Sample cells: first 50 rows, first 20 cols - get formulas and values
        cells_sample = []
        for row in ws.iter_rows(min_row=1, max_row=min(50, ws.max_row), max_col=min(20, ws.max_column), values_only=False):
            for cell in row:
                if cell.value is not None or cell.comment:
                    cells_sample.append({
                        "coordinate": cell.coordinate,
                        "value": cell.value,
                        "formula": cell.value if isinstance(cell.value, str) and cell.value.startswith('=') else None,
                        "data_type": cell.data_type,
                        "number_format": cell.number_format,
                        "has_comment": cell.comment is not None,
                        "comment_text": cell.comment.text if cell.comment else None,
                        "style": {
                            "font": str(cell.font),
                            "fill": str(cell.fill),
                            "alignment": str(cell.alignment),
                            "border": str(cell.border),
                            "number_format": cell.number_format,
                        }
                    })
        sheet_info["cells_sample"] = cells_sample[:200]  # Limit sample size
        
        # Find all formulas
        formulas = []
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column, values_only=False):
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith('='):
                    formulas.append({
                        "coordinate": cell.coordinate,
                        "formula": cell.value,
                        "value": cell.value,
                    })
        sheet_info["formulas_count"] = len(formulas)
        sheet_info["formulas_sample"] = formulas[:100]
        
        # Named ranges in this sheet
        named_ranges = []
        for nr in wb.defined_names.values():
            if nr.attr_text and f"'{sheet_name}'" in nr.attr_text:
                named_ranges.append({
                    "name": nr.name,
                    "ref": nr.attr_text,
                })
        sheet_info["named_ranges"] = named_ranges
        
        inventory["sheets"][sheet_name] = sheet_info
    
    # Workbook-level named ranges
    wb_named_ranges = []
    for nr in wb.defined_names.values():
        wb_named_ranges.append({
            "name": nr.name,
            "ref": nr.attr_text,
            "comment": nr.comment,
        })
    inventory["workbook"]["named_ranges"] = wb_named_ranges
    
    # VBA modules if present
    if wb.vba_archive:
        vba_modules = []
        for name in wb.vba_archive.namelist():
            if name.endswith('.cls') or name.endswith('.bas') or name.endswith('.frm'):
                vba_modules.append(name)
        inventory["workbook"]["vba_modules"] = vba_modules
    
    return inventory

def main():
    print(f"Inspecting: {FILE_PATH}")
    inventory = inspect_workbook(FILE_PATH)
    
    # Save JSON
    json_path = REPORTS_DIR / "workbook_inventory.json"
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(inventory, f, indent=2, ensure_ascii=False, default=str)
    print(f"JSON saved to: {json_path}")
    
    # Generate markdown report
    md_path = REPORTS_DIR / "workbook_inventory.md"
    with open(md_path, 'w', encoding='utf-8') as f:
        f.write(generate_markdown_report(inventory))
    print(f"Markdown saved to: {md_path}")

def generate_markdown_report(inv):
    lines = []
    lines.append("# Workbook Inventory Report")
    lines.append(f"\n**File:** {inv['file_info']['path']}")
    lines.append(f"**Size:** {inv['file_info']['size_bytes']:,} bytes")
    lines.append(f"**Extension:** {inv['file_info']['extension']}")
    lines.append(f"**SHA256:** {inv['file_info']['hashes']['sha256']}")
    lines.append(f"**MD5:** {inv['file_info']['hashes']['md5']}")
    lines.append(f"**Inspection time:** {inv['file_info']['inspection_time']}")
    
    wb = inv['workbook']
    lines.append(f"\n## Workbook Properties")
    lines.append(f"- **Sheets:** {wb['sheet_count']}")
    lines.append(f"- **Active sheet:** {wb['active_sheet']}")
    lines.append(f"- **Has VBA:** {wb['has_vba']}")
    if wb['properties']['creator']:
        lines.append(f"- **Creator:** {wb['properties']['creator']}")
    if wb['properties']['last_modified_by']:
        lines.append(f"- **Last modified by:** {wb['properties']['last_modified_by']}")
    if wb['properties']['created']:
        lines.append(f"- **Created:** {wb['properties']['created']}")
    if wb['properties']['modified']:
        lines.append(f"- **Modified:** {wb['properties']['modified']}")
    
    if wb.get('vba_modules'):
        lines.append(f"\n## VBA Modules ({len(wb['vba_modules'])})")
        for m in wb['vba_modules']:
            lines.append(f"- {m}")
    
    if wb.get('named_ranges'):
        lines.append(f"\n## Workbook Named Ranges ({len(wb['named_ranges'])})")
        for nr in wb['named_ranges']:
            lines.append(f"- **{nr['name']}**: `{nr['ref']}`")
    
    lines.append(f"\n## Sheets Summary")
    for name, sheet in inv['sheets'].items():
        lines.append(f"\n### Sheet: `{name}`")
        lines.append(f"- **State:** {sheet['sheet_state']}")
        lines.append(f"- **Dimensions:** {sheet['dimensions']}")
        lines.append(f"- **Max row:** {sheet['max_row']}")
        lines.append(f"- **Max col:** {sheet['max_column']}")
        lines.append(f"- **Merged cells:** {len(sheet['merged_cells'])}")
        lines.append(f"- **Conditional formats:** {sheet['conditional_formats']}")
        lines.append(f"- **Data validations:** {sheet['data_validations']}")
        lines.append(f"- **Tables:** {len(sheet['tables'])}")
        lines.append(f"- **Charts:** {sheet['charts']}")
        lines.append(f"- **Images:** {sheet['images']}")
        lines.append(f"- **Drawings:** {sheet['drawings']}")
        lines.append(f"- **Comments:** {sheet['comments']}")
        lines.append(f"- **Hyperlinks:** {sheet['hyperlinks']}")
        lines.append(f"- **Auto filter:** {sheet['auto_filter'] or 'None'}")
        lines.append(f"- **Freeze panes:** {sheet['freeze_panes'] or 'None'}")
        lines.append(f"- **Print area:** {sheet['print_area'] or 'None'}")
        lines.append(f"- **Formulas count:** {sheet['formulas_count']}")
        
        if sheet['named_ranges']:
            lines.append(f"- **Sheet named ranges:**")
            for nr in sheet['named_ranges']:
                lines.append(f"  - **{nr['name']}**: `{nr['ref']}`")
        
        if sheet['merged_cells']:
            lines.append(f"- **Merged cells:** {', '.join(sheet['merged_cells'][:10])}{'...' if len(sheet['merged_cells']) > 10 else ''}")
        
        if sheet['tables']:
            lines.append(f"- **Tables:** {', '.join(sheet['tables'])}")
    
    lines.append(f"\n## Detailed Sheet Analysis")
    for name, sheet in inv['sheets'].items():
        lines.append(f"\n### Sheet: `{name}` - Detailed Cells Sample (first 200 non-empty)")
        if sheet['cells_sample']:
            lines.append("| Cell | Value | Formula | Data Type | Number Format | Comment |")
            lines.append("|------|-------|---------|-----------|---------------|---------|")
            for cell in sheet['cells_sample'][:50]:
                val = str(cell['value'])[:80] if cell['value'] is not None else ""
                formula = cell['formula'][:80] if cell['formula'] else ""
                comment = cell['comment_text'][:50] if cell['comment_text'] else ""
                lines.append(f"| {cell['coordinate']} | {val} | {formula} | {cell['data_type']} | {cell['number_format']} | {comment} |")
        else:
            lines.append("*No non-empty cells in sample range*")
        
        if sheet['formulas_sample']:
            lines.append(f"\n#### Formulas Sample (first 50)")
            lines.append("| Cell | Formula |")
            lines.append("|------|---------|")
            for f in sheet['formulas_sample'][:50]:
                lines.append(f"| {f['coordinate']} | `{f['formula']}` |")
    
    return "\n".join(lines)

if __name__ == "__main__":
    main()