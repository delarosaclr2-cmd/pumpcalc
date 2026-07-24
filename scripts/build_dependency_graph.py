import openpyxl
import csv
import re
from pathlib import Path

FILE_PATH = Path(r"C:\PUMPCALC\original\KEETP-60-DM-008 - HOJA DE ESPECIFICACIÓN BOMBA 005PU001 REV C (1).xlsm")
REPORTS_DIR = Path(r"C:\PUMPCALC\reports")

def build_dependency_graph():
    wb = openpyxl.load_workbook(FILE_PATH, read_only=False, keep_vba=True, data_only=False)
    
    # Build cell -> formula mapping
    cell_formulas = {}
    cell_refs = {}  # cell -> set of cells it references
    reverse_refs = {}  # cell -> set of cells that reference it
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column, values_only=False):
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    coord = f"{sheet_name}!{cell.coordinate}"
                    formula = cell.value
                    cell_formulas[coord] = formula
                    
                    # Extract references
                    refs = extract_references(formula, sheet_name)
                    cell_refs[coord] = refs
                    
                    for ref in refs:
                        if ref not in reverse_refs:
                            reverse_refs[ref] = set()
                        reverse_refs[ref].add(coord)
    
    # Identify key sheets and their roles
    key_sheets = [
        'VELOCIDADES RECOMENDADAS',
        'ESPECIFICACIÓN DE TUBERIA',
        'CAIDA PRESION DE TUBERIA',
        'TABLA DE ACCESORIOS SUCCION',
        'TABLA DE ACCESORIOS DESCARGA',
        'RAMALES',
        'CALCULO DE BOMBA',
        '005PU001',
        'RESUMEN PARA PDF',
        'REPORTE GENERAL',
        'REGISTROS'
    ]
    
    # Build sheet-level dependency graph
    sheet_deps = {}
    for sheet in key_sheets:
        sheet_deps[sheet] = {"inputs_from": set(), "outputs_to": set(), "type": ""}
    
    # Classify sheet types
    sheet_types = {
        'VELOCIDADES RECOMENDADAS': 'REFERENCE_TABLE',
        'ESPECIFICACIÓN DE TUBERIA': 'REFERENCE_TABLE',
        'CAIDA PRESION DE TUBERIA': 'CALCULATION_SHEET',
        'TABLA DE ACCESORIOS SUCCION': 'CALCULATION_SHEET',
        'TABLA DE ACCESORIOS DESCARGA': 'CALCULATION_SHEET',
        'RAMALES': 'CALCULATION_SHEET',
        'CALCULO DE BOMBA': 'MAIN_CALCULATION',
        '005PU001': 'OUTPUT_SPEC_SHEET',
        'RESUMEN PARA PDF': 'OUTPUT_REPORT',
        'REPORTE GENERAL': 'OUTPUT_REPORT',
        'REGISTROS': 'LOG_SHEET'
    }
    
    for sheet, stype in sheet_types.items():
        if sheet in sheet_deps:
            sheet_deps[sheet]["type"] = stype
    
    # Find cross-sheet dependencies
    for coord, refs in cell_refs.items():
        source_sheet = coord.split('!')[0]
        for ref in refs:
            target_sheet = ref.split('!')[0] if '!' in ref else source_sheet
            if source_sheet != target_sheet:
                if source_sheet in sheet_deps and target_sheet in sheet_deps:
                    sheet_deps[source_sheet]["inputs_from"].add(target_sheet)
                    sheet_deps[target_sheet]["outputs_to"].add(source_sheet)
    
    # Save edge list
    edges = []
    for coord, refs in cell_refs.items():
        for ref in refs:
            edges.append({
                "source": coord,
                "target": ref,
                "formula": cell_formulas.get(coord, "")
            })
    
    csv_path = REPORTS_DIR / "dependency_edges.csv"
    with open(csv_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=["source", "target", "formula"])
        writer.writeheader()
        writer.writerows(edges)
    
    # Generate markdown report
    md_path = REPORTS_DIR / "dependency_graph.md"
    with open(md_path, 'w', encoding='utf-8') as f:
        f.write("# Dependency Graph\n\n")
        f.write(f"**Total formulas:** {len(cell_formulas)}\n")
        f.write(f"**Total dependency edges:** {len(edges)}\n\n")
        
        f.write("## Sheet-Level Dependencies\n\n")
        for sheet in key_sheets:
            if sheet not in sheet_deps:
                continue
            dep = sheet_deps[sheet]
            f.write(f"### {sheet} ({dep['type']})\n")
            f.write(f"- **Inputs from:** {', '.join(sorted(dep['inputs_from'])) if dep['inputs_from'] else 'None'}\n")
            f.write(f"- **Outputs to:** {', '.join(sorted(dep['outputs_to'])) if dep['outputs_to'] else 'None'}\n\n")
        
        f.write("## Cell-Level Dependencies (Key Calculation Cells)\n\n")
        
        # Focus on key calculation cells in CALCULO DE BOMBA
        key_cells = []
        for coord in cell_formulas:
            sheet = coord.split('!')[0]
            if sheet in ['CALCULO DE BOMBA', 'CAIDA PRESION DE TUBERIA', 'RAMALES', 'TABLA DE ACCESORIOS SUCCION', 'TABLA DE ACCESORIOS DESCARGA']:
                key_cells.append(coord)
        
        for coord in sorted(key_cells)[:100]:
            formula = cell_formulas.get(coord, "")
            refs = cell_refs.get(coord, set())
            refd_by = reverse_refs.get(coord, set())
            f.write(f"### {coord}\n")
            f.write(f"- **Formula:** `{formula}`\n")
            f.write(f"- **References:** {', '.join(sorted(refs)) if refs else 'None'}\n")
            f.write(f"- **Referenced by:** {', '.join(sorted(refd_by)) if refd_by else 'None'}\n\n")
    
    return cell_formulas, cell_refs, reverse_refs, sheet_deps

def extract_references(formula, current_sheet):
    """Extract cell references from a formula"""
    refs = set()
    
    # External references: 'Sheet Name'!CellRef or 'Sheet Name'!RangeRef
    ext_refs = re.findall(r"'([^']+)'!([A-Z]+\$?\d+(?::[A-Z]+\$?\d+)?)", formula)
    for sheet, cell_ref in ext_refs:
        refs.add(f"{sheet}!{cell_ref}")
    
    # Also match without quotes: SheetName!CellRef
    ext_refs2 = re.findall(r"([A-Za-z0-9_]+)!([A-Z]+\$?\d+(?::[A-Z]+\$?\d+)?)", formula)
    for sheet, cell_ref in ext_refs2:
        if sheet != current_sheet:  # Only cross-sheet
            refs.add(f"{sheet}!{cell_ref}")
    
    # Internal references (same sheet): CellRef or RangeRef
    # Match A1, $A$1, A1:B10, $A$1:$B$10 patterns
    int_refs = re.findall(r"(?<!')(?<![A-Za-z0-9_])([A-Z]{1,3}\$?\d+(?::[A-Z]{1,3}\$?\d+)?)(?![A-Za-z0-9_])", formula)
    for cell_ref in int_refs:
        # Skip if it looks like a function name (e.g., SUM, IF)
        if cell_ref.upper() not in ['SUM', 'IF', 'AND', 'OR', 'NOT', 'PI', 'TRUE', 'FALSE', 'VLOOKUP', 'HLOOKUP', 'INDEX', 'MATCH', 'ABS', 'ROUND', 'SQRT', 'POWER', 'EXP', 'LN', 'LOG', 'LOG10', 'SIN', 'COS', 'TAN', 'ASIN', 'ACOS', 'ATAN', 'RADIANS', 'DEGREES']:
            refs.add(f"{current_sheet}!{cell_ref}")
    
    return refs

if __name__ == "__main__":
    cell_formulas, cell_refs, reverse_refs, sheet_deps = build_dependency_graph()
    print(f"Built dependency graph: {len(cell_formulas)} formulas, {sum(len(v) for v in cell_refs.values())} edges")