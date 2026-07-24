import openpyxl
wb = openpyxl.load_workbook(r'C:\PUMPCALC\original\KEETP-60-DM-008 - HOJA DE ESPECIFICACIÓN BOMBA 005PU001 REV C (1).xlsm', data_only=False, keep_vba=True)
ws = wb['VELOCIDADES RECOMENDADAS']

print('FRICCION table (T4:V23):')
for row in range(4, 24):
    vals = []
    for col in [20, 21, 22]:
        cell = ws.cell(row=row, column=col)
        vals.append(cell.value)
    if any(v is not None for v in vals):
        print(f'  Row {row}: {vals}')

print()
print('DIAMETRO table (T4:W23):')
for row in range(4, 24):
    vals = []
    for col in [20, 21, 22, 23]:
        cell = ws.cell(row=row, column=col)
        vals.append(cell.value)
    if any(v is not None for v in vals):
        print(f'  Row {row}: {vals}')

print()
print('RUGOSIDAD table (M26:O31):')
for row in range(26, 32):
    vals = []
    for col in [13, 14, 15]:
        cell = ws.cell(row=row, column=col)
        vals.append(cell.value)
    if any(v is not None for v in vals):
        print(f'  Row {row}: {vals}')

print()
print('OUTPIPES table (B4:I23):')
for row in range(4, 24):
    vals = []
    for col in [2, 3, 4, 5, 6, 7, 8, 9]:
        cell = ws.cell(row=row, column=col)
        vals.append(cell.value)
    if any(v is not None for v in vals):
        print(f'  Row {row}: {vals}')