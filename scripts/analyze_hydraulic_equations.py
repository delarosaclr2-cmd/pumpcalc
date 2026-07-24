import openpyxl
import csv
import re
from pathlib import Path

FILE_PATH = Path(r"C:\PUMPCALC\original\KEETP-60-DM-008 - HOJA DE ESPECIFICACIÓN BOMBA 005PU001 REV C (1).xlsm")
REPORTS_DIR = Path(r"C:\PUMPCALC\reports")

def analyze_hydraulic_equations():
    wb = openpyxl.load_workbook(FILE_PATH, read_only=False, keep_vba=True, data_only=False)
    
    equations = []
    constants = []
    
    # Known hydraulic formulas to identify
    known_constants = {
        14.7: "Standard atmospheric pressure (psi)",
        2.31: "psi to ft water conversion (1 psi = 2.31 ft H2O @ SG=1)",
        2.3071: "Darcy-Weisbach constant for ft head loss (imperial units)",
        50.6: "Reynolds number constant for imperial units (Q in GPM, D in inches, ρ in lb/ft³, μ in cP)",
        3960: "Hydraulic horsepower constant (Q*TDH*SG/3960 = HP)",
        0.7456: "HP to kW conversion (1 HP = 0.7456 kW)",
        5252: "Torque constant (HP * 5252 / RPM = lb-ft)",
        1700: "Motor speed constant (appears to be rated RPM)",
        32.4: "Velocity head constant (V²/(2g) in imperial, g=32.174 ft/s², 2g=64.348, V in ft/s, V²/64.348 = V²/32.174/2... checking: 32.4 is approximately 2g/2? Actually V²/(2g) with g=32.2 -> 64.4, half is 32.2, close to 32.4)",
        448.8309: "GPM to ft³/s conversion (1 GPM = 0.002228 ft³/s, 1/0.002228 = 448.83)",
        3.281: "Meters to feet conversion",
        0.3048: "Feet to meters conversion",
        3.7854: "GPM to LPM conversion (1 GPM = 3.78541 LPM)",
    }
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column, values_only=False):
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    formula = cell.value
                    
                    # Check for known constants in formula
                    for const_val, const_desc in known_constants.items():
                        if str(const_val) in formula:
                            constants.append({
                                "sheet": sheet_name,
                                "cell": cell.coordinate,
                                "formula": formula,
                                "constant_value": const_val,
                                "constant_description": const_desc,
                                "verification_status": "KNOWN"
                            })
                    
                    # Find other numeric constants
                    nums = re.findall(r'(?<![A-Z])(\d+\.?\d*)(?![A-Z])', formula)
                    for num in nums:
                        f = float(num)
                        if f not in known_constants and f not in [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 12, 100, 1000]:
                            # Check if it's a column/row reference
                            if not re.search(rf'[A-Z]\${num}', formula) and not re.search(rf'{num}:', formula):
                                constants.append({
                                    "sheet": sheet_name,
                                    "cell": cell.coordinate,
                                    "formula": formula,
                                    "constant_value": f,
                                    "constant_description": "UNVERIFIED - needs analysis",
                                    "verification_status": "UNVERIFIED"
                                })
    
    # Deduplicate constants
    seen = set()
    unique_constants = []
    for c in constants:
        key = (c['sheet'], c['cell'], c['constant_value'])
        if key not in seen:
            seen.add(key)
            unique_constants.append(c)
    
    # Save constants CSV
    csv_path = REPORTS_DIR / "constants_audit.csv"
    with open(csv_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=unique_constants[0].keys())
        writer.writeheader()
        writer.writerows(unique_constants)
    
    # Now build the hydraulic equations report
    hydraulic_eqs = []
    
    # Define key equations by cell
    key_equations = {
        # CAIDA PRESION DE TUBERIA
        'CAIDA PRESION DE TUBERIA!G8': {
            'variable': 'Diameter (inches) - Discharge',
            'formula': '=G7*(G5/G6)^0.5',
            'equation': 'D = K * sqrt(Q/v)',
            'description': 'Pipe diameter calculation from flow and velocity',
            'units': 'G5=GPM, G6=ft/s, G7=const, D=inches',
            'source': 'Empirical sizing formula'
        },
        'CAIDA PRESION DE TUBERIA!V8': {
            'variable': 'Diameter (inches) - Suction',
            'formula': '=V7*(V5/V6)^0.5',
            'equation': 'D = K * sqrt(Q/v)',
            'description': 'Pipe diameter calculation from flow and velocity',
            'units': 'V5=GPM, V6=ft/s, V7=const, D=inches',
            'source': 'Empirical sizing formula'
        },
        'CAIDA PRESION DE TUBERIA!G11': {
            'variable': 'Reynolds Number - Discharge',
            'formula': '=50.6*G5*G9/(G8*G10)',
            'equation': 'Re = 50.6 * Q * ρ / (D * μ)',
            'description': 'Reynolds number for pipe flow (imperial units)',
            'units': 'Q=GPM, ρ=lb/ft³, D=inches, μ=cP',
            'source': 'Imperial Reynolds formula derivation'
        },
        'CAIDA PRESION DE TUBERIA!V11': {
            'variable': 'Reynolds Number - Suction',
            'formula': '=50.6*V5*V9/(V8*V10)',
            'equation': 'Re = 50.6 * Q * ρ / (D * μ)',
            'description': 'Reynolds number for pipe flow (imperial units)',
            'units': 'Q=GPM, ρ=lb/ft³, D=inches, μ=cP',
            'source': 'Imperial Reynolds formula derivation'
        },
        'CAIDA PRESION DE TUBERIA!G15': {
            'variable': 'Relative Roughness - Discharge',
            'formula': '=G14/(G12/12)',
            'equation': 'ε/D = ε_abs / D_ft',
            'description': 'Relative roughness for Moody chart',
            'units': 'G14=ft (absolute roughness), G12=inches (nominal diameter)',
            'source': 'Standard definition'
        },
        'CAIDA PRESION DE TUBERIA!V15': {
            'variable': 'Relative Roughness - Suction',
            'formula': '=V14/(V12/12)',
            'equation': 'ε/D = ε_abs / D_ft',
            'description': 'Relative roughness for Moody chart',
            'units': 'V14=ft (absolute roughness), V12=inches (nominal diameter)',
            'source': 'Standard definition'
        },
        'CAIDA PRESION DE TUBERIA!V16': {
            'variable': 'Friction Factor (Laminar) - Suction',
            'formula': '= 64/V11',
            'equation': 'f = 64/Re',
            'description': 'Laminar flow friction factor (Hagen-Poiseuille)',
            'units': 'Re = Reynolds number',
            'source': 'Theoretical laminar flow'
        },
        'CAIDA PRESION DE TUBERIA!G19': {
            'variable': 'Friction Head Loss (ft/ft) - Discharge',
            'formula': '=(((G17*G16*G9*(G5^2))/(G8^5))*2.3071)*G18',
            'equation': 'hf/L = f * (L/D) * (V²/2g) -> converted to ft/ft using Q',
            'description': 'Darcy-Weisbach head loss per unit length (imperial)',
            'units': 'G17=f, G16=?, G9=lb/ft³, G5=GPM, G8=inches, 2.3071=const, G18=length factor',
            'source': 'Darcy-Weisbach with imperial conversions'
        },
        'CAIDA PRESION DE TUBERIA!V19': {
            'variable': 'Friction Head Loss (ft/ft) - Suction',
            'formula': '=(((V17*V16*V9*(V5^2))/(V8^5))*2.3071)*V18',
            'equation': 'hf/L = f * (L/D) * (V²/2g) -> converted to ft/ft using Q',
            'description': 'Darcy-Weisbach head loss per unit length (imperial)',
            'units': 'V17=f, V16=?, V9=lb/ft³, V5=GPM, V8=inches, 2.3071=const, V18=length factor',
            'source': 'Darcy-Weisbach with imperial conversions'
        },
        # CALCULO DE BOMBA
        'CALCULO DE BOMBA!C9': {
            'variable': 'Static Suction Head (ft)',
            'formula': '=500/304.8',
            'equation': 'Hs = elevation_diff_m / 0.3048',
            'description': 'Static suction head from elevation difference',
            'units': '500 = mm, 304.8 = mm/ft, result = ft',
            'source': 'Geometry conversion'
        },
        'CALCULO DE BOMBA!C12': {
            'variable': 'Suction Pipe Length (ft)',
            'formula': '=2.12*3.281',
            'equation': 'L_ft = L_m * 3.281',
            'description': 'Suction pipe length conversion from meters to feet',
            'units': '2.12 = meters, 3.281 = ft/m, result = ft',
            'source': 'Unit conversion'
        },
        'CALCULO DE BOMBA!C13': {
            'variable': 'Suction Friction Loss per ft (ft/ft)',
            'formula': "='CAIDA PRESION DE TUBERIA'!V19",
            'equation': 'hf_suction_per_ft = from suction pipe calc',
            'description': 'Suction line friction loss per foot',
            'units': 'ft/ft',
            'source': 'From CAIDA PRESION DE TUBERIA suction calc'
        },
        'CALCULO DE BOMBA!C14': {
            'variable': 'Total Suction Pipe Friction Loss (ft)',
            'formula': '=C12*C13',
            'equation': 'Hf_suction = L_suction * hf_per_ft',
            'description': 'Total suction pipe friction loss',
            'units': 'ft',
            'source': 'Basic multiplication'
        },
        'CALCULO DE BOMBA!E14': {
            'variable': 'NPSH Available (ft)',
            'formula': '=((C8+E8)*(2.31/E11))+C9-C11-C14-E9',
            'equation': 'NPSHa = (Patm + Pvessel) * 2.31/SG + Hs - Hf_acc - Hf_pipe - Pv',
            'description': 'Net Positive Suction Head Available',
            'units': 'C8=psia, E8=psig, 2.31=psi->ft, E11=SG, C9=ft, C11=ft, C14=ft, E9=ft',
            'source': 'Standard NPSH formula'
        },
        'CALCULO DE BOMBA!E20': {
            'variable': 'Hydraulic Power (HP)',
            'formula': '=(E4*C28*E11)/3960',
            'equation': 'P_hyd = Q * TDH * SG / 3960',
            'description': 'Hydraulic horsepower',
            'units': 'Q=GPM, TDH=ft, SG=specific gravity, 3960=constant, result=HP',
            'source': 'Standard pump power formula'
        },
        'CALCULO DE BOMBA!E21': {
            'variable': 'Brake Power (HP)',
            'formula': '=E20/C22',
            'equation': 'P_brake = P_hyd / η_pump',
            'description': 'Power at pump shaft',
            'units': 'E20=hydraulic HP, C22=pump efficiency, result=HP',
            'source': 'Pump power definition'
        },
        'CALCULO DE BOMBA!E22': {
            'variable': 'Brake Power (kW)',
            'formula': '=E21*0.7456',
            'equation': 'P_kW = P_HP * 0.7456',
            'description': 'Brake power conversion to kW',
            'units': 'HP to kW',
            'source': 'Unit conversion'
        },
        'CALCULO DE BOMBA!E23': {
            'variable': 'Torque (lb-ft)',
            'formula': '=(E21*5252)/1700',
            'equation': 'T = HP * 5252 / RPM',
            'description': 'Shaft torque at rated speed',
            'units': 'E21=HP, 5252=constant, 1700=RPM, result=lb-ft',
            'source': 'Torque-power-speed relationship'
        },
        'CALCULO DE BOMBA!E24': {
            'variable': 'TDH (meters)',
            'formula': '=C28*0.3048',
            'equation': 'TDH_m = TDH_ft * 0.3048',
            'description': 'Total Dynamic Head in meters',
            'units': 'ft to m conversion',
            'source': 'Unit conversion'
        },
        'CALCULO DE BOMBA!E25': {
            'variable': 'Flow (LPM)',
            'formula': '=E4*3.7854',
            'equation': 'Q_LPM = Q_GPM * 3.7854',
            'description': 'Flow rate conversion to LPM',
            'units': 'GPM to LPM',
            'source': 'Unit conversion'
        },
        'CALCULO DE BOMBA!E27': {
            'variable': 'Specific Speed (Ns)',
            'formula': '=(C29*(E4^0.5))/(E24^0.75)',
            'equation': 'Ns = N * sqrt(Q) / H^0.75',
            'description': 'Pump specific speed (imperial units)',
            'units': 'N=RPM, Q=GPM, H=ft, result=dimensionless (US units)',
            'source': 'Specific speed definition'
        },
        'CALCULO DE BOMBA!C28': {
            'variable': 'Total Dynamic Head (ft)',
            'formula': '=C11+C14+C21+C24+C26',
            'equation': 'TDH = Hf_suc_acc + Hf_suc_pipe + H_static + Hf_dis_acc + Hf_dis_pipe',
            'description': 'Total Dynamic Head = sum of all head losses + static head',
            'units': 'All terms in ft',
            'source': 'System head summation'
        },
        # RAMALES
        'RAMALES!D10': {
            'variable': 'Velocity (ft/s)',
            'formula': '=(4*(D9/448.8309))/(PI()*(D8/12)*(D8/12))',
            'equation': 'V = 4Q / (πD²)',
            'description': 'Flow velocity in pipe',
            'units': 'D9=GPM, 448.8309=GPM->ft³/s, D8=inches, result=ft/s',
            'source': 'Continuity equation'
        },
        'RAMALES!D12': {
            'variable': 'Discharge Pipe Friction Loss per ft (ft/ft)',
            'formula': '=(((\'CAIDA PRESION DE TUBERIA\'!$G$17*\'CAIDA PRESION DE TUBERIA\'!$G$16*\'CAIDA PRESION DE TUBERIA\'!$G$9*(D9^2))/((\'CAIDA PRESION DE TUBERIA\'!$G$7*(D9/D10)^0.5)^5))*2.3071)*\'CAIDA PRESION DE TUBERIA\'!$G$18',
            'equation': 'hf/L = f * (V²/2g) / D',
            'description': 'Discharge pipe friction loss per unit length',
            'units': 'Complex - uses friction factor, velocity, density from other sheet',
            'source': 'Darcy-Weisbach via reference sheet'
        },
        # RESUMEN PARA PDF
        'RESUMEN PARA PDF!B28': {
            'variable': 'Total Dynamic Head (ft)',
            'formula': '=(B21+C23+C24+C25+C26)*1',
            'equation': 'TDH = H_static + Hf_suc_acc + Hf_suc_pipe + Hf_dis_acc + Hf_dis_pipe',
            'description': 'Total Dynamic Head from component summation',
            'units': 'All terms in ft',
            'source': 'System head curve summation'
        },
        'RESUMEN PARA PDF!D28': {
            'variable': 'TDH (psi)',
            'formula': '=B28/2.31',
            'equation': 'P_psi = TDH_ft / 2.31',
            'description': 'Head to pressure conversion',
            'units': 'ft to psi (assuming SG=1)',
            'source': 'Standard conversion'
        },
        # TABLA DE ACCESORIOS
        'TABLA DE ACCESORIOS DESCARGA!I7': {
            'variable': 'Accessory Equivalent Length (ft)',
            'formula': '=((D7*F7)*($H$2^2)/(32.4*2))*H7',
            'equation': 'Leq = K * V² / (2g) * (1/D) * count? or Leq = f * L/D...',
            'description': 'Equivalent length of fitting (Crane method)',
            'units': 'D7=ft (from table), F7=K factor, H2=V ft/s, 32.4=2g?, H7=quantity',
            'source': 'Crane TP-410 method'
        },
        'TABLA DE ACCESORIOS DESCARGA!U41': {
            'variable': 'Sum of accessory losses (ft)',
            'formula': '=(SUM(U7:U40))*2.31',
            'equation': 'Total_accessory_head = sum * 2.31',
            'description': 'Accessory losses converted to head',
            'units': 'U7:U40=psi?, *2.31 -> ft',
            'source': 'Pressure to head conversion'
        },
    }
    
    # Save hydraulic equations markdown
    md_path = REPORTS_DIR / "hydraulic_equations.md"
    with open(md_path, 'w', encoding='utf-8') as f:
        f.write("# Hydraulic Equations Inventory\n\n")
        f.write("This document catalogs all identified hydraulic calculations in the workbook.\n\n")
        
        # Group by sheet
        by_sheet = {}
        for cell_ref, eq in key_equations.items():
            sheet = cell_ref.split('!')[0]
            if sheet not in by_sheet:
                by_sheet[sheet] = []
            by_sheet[sheet].append((cell_ref, eq))
        
        for sheet in sorted(by_sheet.keys()):
            f.write(f"## {sheet}\n\n")
            for cell_ref, eq in by_sheet[sheet]:
                f.write(f"### {cell_ref}\n")
                f.write(f"- **Variable:** {eq['variable']}\n")
                f.write(f"- **Excel Formula:** `{eq['formula']}`\n")
                f.write(f"- **Engineering Equation:** {eq['equation']}\n")
                f.write(f"- **Description:** {eq['description']}\n")
                f.write(f"- **Units:** {eq['units']}\n")
                f.write(f"- **Source/Reference:** {eq['source']}\n")
                f.write(f"- **Verification Status:** PENDING\n\n")
        
        # Add constants section
        f.write("---\n\n## Numerical Constants Audit\n\n")
        f.write("| Sheet | Cell | Constant | Description | Status |\n")
        f.write("|-------|------|----------|-------------|--------|\n")
        for c in unique_constants:
            f.write(f"| {c['sheet']} | {c['cell']} | {c['constant_value']} | {c['constant_description']} | {c['verification_status']} |\n")
    
    # Save constants CSV (already done)
    # Save equations CSV
    eq_csv_path = REPORTS_DIR / "calculation_variables.csv"
    eq_rows = []
    for cell_ref, eq in key_equations.items():
        eq_rows.append({
            "cell_ref": cell_ref,
            **eq
        })
    with open(eq_csv_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=eq_rows[0].keys())
        writer.writeheader()
        writer.writerows(eq_rows)
    
    return key_equations, unique_constants

if __name__ == "__main__":
    eqs, consts = analyze_hydraulic_equations()
    print(f"Identified {len(eqs)} key hydraulic equations")
    print(f"Found {len(consts)} numerical constants")