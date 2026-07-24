"""
Excel reader module - Read workbook and extract data without modification.
"""
import openpyxl
from typing import Dict, List, Optional, Any
from dataclasses import dataclass
from pathlib import Path


@dataclass
class CellData:
    """Cell data with full context."""
    sheet: str
    cell: str
    value: Any
    formula: Optional[str]
    number_format: str
    data_type: str
    label: Optional[str] = None  # text in adjacent cell
    unit: Optional[str] = None   # unit from label or header


class ExcelReader:
    """Read Excel workbook without modifying it."""
    
    def __init__(self, filepath: str):
        self.filepath = Path(filepath)
        self.wb = openpyxl.load_workbook(
            filepath, 
            read_only=False, 
            keep_vba=True, 
            data_only=False
        )
    
    def get_all_formulas(self) -> List[CellData]:
        """Extract all formulas with context."""
        formulas = []
        for sheet_name in self.wb.sheetnames:
            ws = self.wb[sheet_name]
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column, values_only=False):
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        # Get label from left cell
                        label = None
                        if cell.column > 1:
                            label_cell = ws.cell(row=cell.row, column=cell.column-1)
                            if label_cell.value:
                                label = str(label_cell.value)
                        
                        formulas.append(CellData(
                            sheet=sheet_name,
                            cell=cell.coordinate,
                            value=cell.value,
                            formula=cell.value,
                            number_format=cell.number_format,
                            data_type=cell.data_type,
                            label=label
                        ))
        return formulas
    
    def get_cell_value(self, sheet: str, cell: str) -> Any:
        """Get cell value (cached or formula)."""
        ws = self.wb[sheet]
        return ws[cell].value
    
    def get_sheet_names(self) -> List[str]:
        return self.wb.sheetnames
    
    def get_named_ranges(self) -> Dict[str, str]:
        """Get all named ranges."""
        ranges = {}
        for nr in self.wb.defined_names.values():
            if nr.attr_text:
                ranges[nr.name] = nr.attr_text
        return ranges
    
    def close(self):
        self.wb.close()


def extract_workbook_inputs(filepath: str) -> Dict[str, Any]:
    """Extract all input values from workbook for the current case."""
    reader = ExcelReader(filepath)
    
    # Key cells from the workbook (from current_case_inputs)
    key_cells = {
        # CAIDA PRESION DE TUBERIA
        'flow_gpm_discharge': ('CAIDA PRESION DE TUBERIA', 'G5'),
        'vel_discharge': ('CAIDA PRESION DE TUBERIA', 'G6'),
        'vel_suction': ('CAIDA PRESION DE TUBERIA', 'V6'),
        'const_k': ('CAIDA PRESION DE TUBERIA', 'G7'),
        'density_discharge': ('CAIDA PRESION DE TUBERIA', 'G9'),
        'viscosity_discharge': ('CAIDA PRESION DE TUBERIA', 'G10'),
        'density_suction': ('CAIDA PRESION DE TUBERIA', 'V9'),
        'viscosity_suction': ('CAIDA PRESION DE TUBERIA', 'V10'),
        'nominal_d_discharge': ('CAIDA PRESION DE TUBERIA', 'G12'),
        'nominal_d_suction': ('CAIDA PRESION DE TUBERIA', 'V12'),
        'roughness_abs_discharge': ('CAIDA PRESION DE TUBERIA', 'G14'),
        'roughness_abs_suction': ('CAIDA PRESION DE TUBERIA', 'V14'),
        'fluid_code_discharge': ('CAIDA PRESION DE TUBERIA', 'A20'),
        'fluid_code_suction': ('CAIDA PRESION DE TUBERIA', 'B20'),
        'roughness_code_discharge': ('CAIDA PRESION DE TUBERIA', 'A21'),
        'roughness_code_suction': ('CAIDA PRESION DE TUBERIA', 'B21'),
        
        # CALCULO DE BOMBA
        'p_atm': ('CALCULO DE BOMBA', 'C8'),
        'p_vessel': ('CALCULO DE BOMBA', 'E8'),
        'static_suction_head_ft': ('CALCULO DE BOMBA', 'C9'),
        'vapor_pressure_ft': ('CALCULO DE BOMBA', 'E9'),
        'suction_fitting_losses': ('CALCULO DE BOMBA', 'C11'),
        'specific_gravity': ('CALCULO DE BOMBA', 'E11'),
        'suction_pipe_length': ('CALCULO DE BOMBA', 'C12'),
        'suction_pipe_loss_per_ft': ('CALCULO DE BOMBA', 'C13'),
        'suction_pipe_losses': ('CALCULO DE BOMBA', 'C14'),
        'npsha': ('CALCULO DE BOMBA', 'E14'),
        'static_discharge_head': ('CALCULO DE BOMBA', 'C20'),
        'pump_efficiency': ('CALCULO DE BOMBA', 'C22'),
        'discharge_fitting_losses': ('CALCULO DE BOMBA', 'C24'),
        'tdh_ft': ('CALCULO DE BOMBA', 'C28'),
        'rpm': ('CALCULO DE BOMBA', 'C29'),
        'fluid_code_pv': ('CALCULO DE BOMBA', 'A32'),
        
        # RAMALES
        'discharge_pipe_length': ('RAMALES', 'D11'),
        'discharge_pipe_diameter': ('RAMALES', 'D8'),
        
        # RESUMEN PARA PDF
        'tdh_ft_resumen': ('RESUMEN PARA PDF', 'B28'),
        'npsha_resumen': ('RESUMEN PARA PDF', 'G25'),
        'power_hp_resumen': ('RESUMEN PARA PDF', 'G29'),
    }
    
    inputs = {}
    for key, (sheet, cell) in key_cells.items():
        try:
            val = reader.get_cell_value(sheet, cell)
            inputs[key] = val
        except Exception as e:
            inputs[key] = f"ERROR: {e}"
    
    reader.close()
    return inputs


if __name__ == '__main__':
    filepath = r"C:\PUMPCALC\original\KEETP-60-DM-008 - HOJA DE ESPECIFICACIÓN BOMBA 005PU001 REV C (1).xlsm"
    inputs = extract_workbook_inputs(filepath)
    
    print("Extracted Workbook Inputs:")
    print("=" * 60)
    for k, v in inputs.items():
        print(f"  {k}: {v}")